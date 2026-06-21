"""
Analiză cheltuieli din subvenție — aplicație Streamlit.

Utilizatorul încarcă fișierul Excel (.xls / .xlsx) folosit pentru raportare.
Aplicația RECALCULEAZĂ singură toate totalurile, pornind DOAR de la rândurile
de tranzacții (cele cu dată reală), ignorând complet rândurile de subtotal/total
din fișier. Astfel rezultatele nu depind de formulele din Excel (care pot fi greșite).

Structura fișierului se recunoaște după coloanele FIXE (Luna, data, Total cheltuită,
Total primită, Venituri în clarificare, Sume transferate, Sold cont), iar categoriile
de cheltuieli sunt deduse automat (tot ce stă între coloana de date și „Total cheltuită"),
deci merge chiar dacă se schimbă categoriile de la an la an.
"""

import io
import re
import datetime
import unicodedata

import numpy as np
import pandas as pd


# ----------------------------------------------------------------------------- #
#  LOGICĂ (pură, testabilă fără Streamlit)
# ----------------------------------------------------------------------------- #

def _norm(s):
    """Normalizează un text: fără diacritice, litere mici, spații colapsate."""
    if s is None or (isinstance(s, float) and np.isnan(s)):
        return ""
    s = unicodedata.normalize("NFKD", str(s))
    s = "".join(c for c in s if not unicodedata.combining(c))
    return re.sub(r"\s+", " ", s.lower()).strip()


def _to_date(v):
    """Întoarce un Timestamp DOAR pentru valori care sunt cu adevărat date
    (datetime). Textele de tip 'IANUARIE' NU sunt considerate date."""
    if isinstance(v, (pd.Timestamp, datetime.datetime, datetime.date)):
        try:
            return pd.Timestamp(v)
        except Exception:
            return pd.NaT
    return pd.NaT


def _to_num(series):
    """Coerce la numeric, tratând textul/golul ca 0."""
    return pd.to_numeric(series, errors="coerce").fillna(0.0)


def read_excel_any(file_like, filename):
    engine = "xlrd" if filename.lower().rsplit(".", 1)[-1] == "xls" else "openpyxl"
    return pd.read_excel(file_like, header=None, engine=engine, dtype=object)


def detect_header_row(raw):
    """Caută rândul de antet (cel cu 'Luna' și 'Sold cont' / 'Total cheltuită')."""
    for i in range(min(15, len(raw))):
        vals = [_norm(v) for v in raw.iloc[i].tolist()]
        has_luna = any(v == "luna" or v.startswith("luna") for v in vals)
        has_sold = any("sold cont" in v for v in vals)
        has_tot = any("total" in v and "cheltuit" in v for v in vals)
        if has_luna and (has_sold or has_tot):
            return i
    return 0


def find_anchor_columns(headers):
    """headers = listă de antete normalizate. Întoarce indecșii coloanelor fixe."""
    def find(pred):
        for idx, h in enumerate(headers):
            if pred(h):
                return idx
        return None

    return {
        "luna": find(lambda h: h == "luna" or h.startswith("luna")),
        "total_cheltuit": find(lambda h: "total" in h and "cheltuit" in h),
        "total_primit": find(lambda h: "total" in h and "primit" in h),
        # IMPORTANT: 'venituri ... clarificare' (AE), nu categoria 'cheltuieli in curs de clarificare'
        "venit_clarificare": find(lambda h: "venitur" in h and "clarificare" in h),
        "transferat": find(lambda h: "transferate" in h or ("sume" in h and "filiale" in h)),
        "sold": find(lambda h: "sold cont" in h or h == "sold"),
        "furnizor": find(lambda h: "furnizor" in h),
        "factura": find(lambda h: "factura" in h or ("nr" in h and "factur" in h)),
    }


def detect_date_col(df, luna_idx, total_cheltuit_idx):
    """Coloana de date = cea cu cele mai multe date reale, între Luna și Total cheltuită."""
    start = (luna_idx + 1) if luna_idx is not None else 0
    end = total_cheltuit_idx if total_cheltuit_idx is not None else df.shape[1]
    best, best_cnt = None, -1
    for c in range(start, max(start, end)):
        cnt = int(df.iloc[:, c].map(lambda v: not pd.isna(_to_date(v))).sum())
        if cnt > best_cnt:
            best_cnt, best = cnt, c
    return best


def _xl_col(idx):
    """Index coloană (0-based) -> literă Excel (0->A, 27->AB, 28->AC)."""
    n, s = idx + 1, ""
    while n > 0:
        n, r = divmod(n - 1, 26)
        s = chr(65 + r) + s
    return s


MONTHS_RO = ["ianuarie", "februarie", "martie", "aprilie", "mai", "iunie",
             "iulie", "august", "septembrie", "octombrie", "noiembrie", "decembrie"]


def _luna_ro(period):
    return f"{MONTHS_RO[period.month - 1]} {period.year}"


def analyze(raw):
    """Analiza completă. Întoarce un dict cu rezultate sau ridică ValueError clar."""
    hdr = detect_header_row(raw)
    headers_norm = [_norm(v) for v in raw.iloc[hdr].tolist()]
    headers_raw = ["" if pd.isna(v) else str(v).strip() for v in raw.iloc[hdr].tolist()]
    cols = find_anchor_columns(headers_norm)

    if cols["total_cheltuit"] is None:
        raise ValueError("Nu am găsit coloana „Total subvenție cheltuită”. "
                         "Verifică antetul fișierului.")

    body_full = raw.iloc[hdr + 1:].reset_index(drop=True)

    # Limita blocului de sinteză: primul rând unde coloana Luna conține „total"
    # (ex. „TOTAL CATEGORIE"). De acolo în jos = sinteză manuală, o excludem.
    boundary, grand = None, None
    if cols["luna"] is not None:
        for i in range(len(body_full)):
            if "total" in _norm(body_full.iloc[i, cols["luna"]]):
                boundary = i
                val = pd.to_numeric(pd.Series([body_full.iloc[i, cols["total_cheltuit"]]]),
                                    errors="coerce").iloc[0]
                if pd.notna(val):
                    grand = {"rand": i + hdr + 2, "total_fisier": float(val)}
                break
    body = body_full.iloc[:boundary] if boundary is not None else body_full

    date_idx = detect_date_col(body, cols["luna"], cols["total_cheltuit"])
    if date_idx is None:
        raise ValueError("Nu am găsit coloana cu date (datele tranzacțiilor).")

    cat_idxs = list(range(date_idx + 1, cols["total_cheltuit"]))
    if not cat_idxs:
        raise ValueError("Nu am găsit nicio coloană de categorie între dată și „Total cheltuită”.")
    cat_names = [headers_raw[i] or f"Coloana {i + 1}" for i in cat_idxs]

    dates_all = body.iloc[:, date_idx].map(_to_date)
    mask_tx = dates_all.notna()
    tx = body[mask_tx]              # păstrăm indexul original (pt. nr. rând corect)
    tx_dates = dates_all[mask_tx]

    # --- sume pe rând ---
    cat_matrix = tx.iloc[:, cat_idxs].apply(_to_num)
    cat_matrix.columns = cat_names
    row_cheltuit = cat_matrix.sum(axis=1)
    primit = _to_num(tx.iloc[:, cols["total_primit"]]) if cols["total_primit"] is not None else pd.Series(0.0, index=tx.index)
    venit = _to_num(tx.iloc[:, cols["venit_clarificare"]]) if cols["venit_clarificare"] is not None else pd.Series(0.0, index=tx.index)
    transferat = _to_num(tx.iloc[:, cols["transferat"]]) if cols["transferat"] is not None else pd.Series(0.0, index=tx.index)
    movement = primit + venit - row_cheltuit - transferat

    # --- sold inițial (preluat din fișier, e dată reală nu formulă) ---
    opening = 0.0
    opening_source = "presupus 0 (nu am găsit coloana Sold cont)"
    if cols["sold"] is not None:
        sold_typed = pd.to_numeric(tx.iloc[:, cols["sold"]], errors="coerce")
        fv = sold_typed.first_valid_index()
        if fv is not None:
            opening = float(sold_typed.loc[fv] - movement.loc[fv])
            opening_source = "preluat din fișier (primul Sold cont)"

    running = opening + movement.cumsum()
    final_sold = float(opening + movement.sum())

    # --- pe categorii ---
    cat_tot = cat_matrix.sum(axis=0)
    cat_tot = cat_tot[cat_tot != 0].sort_values(ascending=False)
    total_cheltuit = float(row_cheltuit.sum())
    cat_table = pd.DataFrame({
        "Categorie": cat_tot.index,
        "Total (lei)": cat_tot.values,
        "% din cheltuieli": (cat_tot.values / total_cheltuit * 100) if total_cheltuit else 0,
    })

    # --- pe luni ---
    period = tx_dates.dt.to_period("M")
    mdf = pd.DataFrame({
        "p": period.values,
        "primit": primit.values, "cheltuit": row_cheltuit.values,
        "venit": venit.values, "transferat": transferat.values,
        "movement": movement.values,
    })
    g = mdf.groupby("p", sort=True).sum(numeric_only=True)
    g["sold_sf_luna"] = opening + g["movement"].cumsum()
    month_table = pd.DataFrame({
        "Luna": [_luna_ro(p) for p in g.index],
        "Primit (lei)": g["primit"].values,
        "Cheltuit (lei)": g["cheltuit"].values,
        "Sold la sfârșit de lună (lei)": g["sold_sf_luna"].values,
    })

    # --- pe zile ---
    ddf = pd.DataFrame({
        "zi": tx_dates.dt.normalize().values,
        "primit": primit.values, "cheltuit": row_cheltuit.values,
        "movement": movement.values,
    })
    gd = ddf.groupby("zi", sort=True).sum(numeric_only=True)
    gd["sold_sf_zi"] = opening + gd["movement"].cumsum()
    day_table = pd.DataFrame({
        "Data": [pd.Timestamp(z).date() for z in gd.index],
        "Primit (lei)": gd["primit"].values,
        "Cheltuit (lei)": gd["cheltuit"].values,
        "Sold la sfârșit de zi (lei)": gd["sold_sf_zi"].values,
    })

    # --- pe furnizor ---
    supplier_table = None
    if cols["furnizor"] is not None:
        sdf = pd.DataFrame({
            "Furnizor": tx.iloc[:, cols["furnizor"]].map(lambda v: "(necompletat)" if pd.isna(v) else str(v).strip()).values,
            "cheltuit": row_cheltuit.values,
        })
        sdf = sdf[sdf["cheltuit"] != 0]
        supplier_table = (sdf.groupby("Furnizor", as_index=False)["cheltuit"].sum()
                          .sort_values("cheltuit", ascending=False)
                          .rename(columns={"cheltuit": "Total cheltuit (lei)"}))

    # --- sold pe timp (pentru grafic) ---
    sold_curve = pd.DataFrame({"Data": tx_dates.values, "Sold (lei)": running.values})
    sold_curve = sold_curve.dropna().sort_values("Data")

    # ======================= VERIFICĂRI =======================
    checks = {}

    # 1) Total pe rând tastat ≠ suma categoriilor
    if cols["total_cheltuit"] is not None:
        typed_ac = pd.to_numeric(tx.iloc[:, cols["total_cheltuit"]], errors="coerce")
        diff = (typed_ac - row_cheltuit)
        bad = diff.abs() > 0.01
        rows = []
        for i in np.where(bad & typed_ac.notna())[0]:
            rows.append({
                "Rând în fișier": int(tx.index[i]) + hdr + 2,
                "Total tastat": round(float(typed_ac.iloc[i]), 2),
                "Suma categoriilor": round(float(row_cheltuit.iloc[i]), 2),
                "Diferență": round(float(diff.iloc[i]), 2),
            })
        checks["total_rand_gresit"] = pd.DataFrame(rows)

    # 2) Tranzacții reale FĂRĂ dată: au furnizor + bani, dar nu au dată -> NU sunt
    #    numărate în totaluri (rândurile de subtotal lunar nu au furnizor, deci sunt excluse aici).
    no_date = body[~mask_tx]
    if len(no_date) and cols["furnizor"] is not None:
        ndc = no_date.iloc[:, cat_idxs].apply(_to_num).sum(axis=1)
        furn = no_date.iloc[:, cols["furnizor"]]

        def _real_supplier(v):
            if pd.isna(v):
                return False
            s = str(v).strip()
            if s == "":
                return False
            return not re.fullmatch(r"[-\d.,\s]+", s)  # nume real conține litere

        has_furn = furn.map(_real_supplier)
        skipped = no_date[(ndc != 0) & has_furn]
        rows = []
        for idx in skipped.index:
            rows.append({
                "Rând în fișier": int(idx) + hdr + 2,
                "Furnizor": str(furn.loc[idx]).strip(),
                "Sumă (lei)": round(float(ndc.loc[idx]), 2),
            })
        checks["orfane"] = pd.DataFrame(rows)

    # 3) Sold din fișier ≠ sold recalculat — rezumat (nu listă lungă)
    if cols["sold"] is not None:
        sold_typed = pd.to_numeric(tx.iloc[:, cols["sold"]], errors="coerce")
        sdiff = (sold_typed - running)
        mism = (sdiff.abs() > 0.01) & sold_typed.notna()
        idxs = np.where(mism)[0]
        if len(idxs):
            first = int(idxs[0])
            last_valid = sold_typed.last_valid_index()
            checks["sold"] = {
                "n": int(len(idxs)),
                "first_row": int(tx.index[first]) + hdr + 2,
                "first_typed": round(float(sold_typed.iloc[first]), 2),
                "first_correct": round(float(running.iloc[first]), 2),
                "final_typed": round(float(sold_typed.loc[last_valid]), 2) if last_valid is not None else None,
                "final_correct": round(float(running.iloc[-1]), 2),
            }
        else:
            checks["sold"] = {"n": 0}

    # ======================= DETALIU TRANZACȚII (pt. potrivirea cu extrasul) =======================
    fact_idx = cols.get("factura")
    furn_idx = cols.get("furnizor")
    typed_total = pd.to_numeric(tx.iloc[:, cols["total_cheltuit"]], errors="coerce")
    amt_for_match = typed_total.where(typed_total.notna() & (typed_total != 0), row_cheltuit)
    sold_idx = cols.get("sold")
    sold_typed_detail = (pd.to_numeric(tx.iloc[:, sold_idx], errors="coerce").values
                         if sold_idx is not None else np.full(len(tx), np.nan))
    tx_detail = pd.DataFrame({
        "Rând": [int(i) + hdr + 2 for i in tx.index],
        "Data": tx_dates.values,
        "Furnizor": (tx.iloc[:, furn_idx].map(lambda v: "" if pd.isna(v) else str(v).strip()).values
                     if furn_idx is not None else ""),
        "Nr. factură": (tx.iloc[:, fact_idx].map(lambda v: "" if pd.isna(v) else str(v).strip()).values
                        if fact_idx is not None else ""),
        "Sumă (lei)": amt_for_match.values,
        "Sold AG (raport)": sold_typed_detail,
    })

    # ======================= CALITATEA DATELOR =======================
    # Considerăm „rând de tranzacție" orice rând din corp care are furnizor real SAU sumă.
    quality = {}
    cat_full = body.iloc[:, cat_idxs].apply(_to_num)
    sum_full = cat_full.sum(axis=1)
    typed_full = pd.to_numeric(body.iloc[:, cols["total_cheltuit"]], errors="coerce")
    amt_full = typed_full.where(typed_full.notna() & (typed_full != 0), sum_full)
    dates_full = body.iloc[:, date_idx].map(_to_date)
    furn_full = (body.iloc[:, furn_idx] if furn_idx is not None else pd.Series([np.nan] * len(body), index=body.index))
    fact_full = (body.iloc[:, fact_idx] if fact_idx is not None else pd.Series([np.nan] * len(body), index=body.index))

    def _has_text(v):
        return not (pd.isna(v) or str(v).strip() == "")

    def _real_name(v):
        if not _has_text(v):
            return False
        return not re.fullmatch(r"[-\d.,\s]+", str(v).strip())

    is_txrow = (amt_full.fillna(0) != 0) | furn_full.map(_real_name)

    # 1) celule lipsă
    miss_rows = []
    for idx in body.index[is_txrow]:
        lipsa = []
        if pd.isna(dates_full.loc[idx]):
            lipsa.append("dată")
        if not _real_name(furn_full.loc[idx]):
            lipsa.append("furnizor")
        if not _has_text(fact_full.loc[idx]):
            lipsa.append("nr. factură")
        if pd.isna(amt_full.loc[idx]) or amt_full.loc[idx] == 0:
            lipsa.append("sumă")
        if lipsa:
            miss_rows.append({
                "Rând în fișier": int(idx) + hdr + 2,
                "Furnizor": "" if not _has_text(furn_full.loc[idx]) else str(furn_full.loc[idx]).strip(),
                "Sumă (lei)": round(float(amt_full.loc[idx]), 2) if pd.notna(amt_full.loc[idx]) else 0.0,
                "Câmpuri lipsă": ", ".join(lipsa),
            })
    quality["lipsuri"] = pd.DataFrame(miss_rows)

    # 2) ani diferiți de anul dominant (raportul ar trebui să fie pe un singur an)
    yrs = tx_dates.dt.year
    year_counts = yrs.value_counts().sort_index()
    quality["year_counts"] = year_counts
    dominant_year = int(year_counts.idxmax()) if len(year_counts) else None
    quality["dominant_year"] = dominant_year
    wrong_year = []
    if dominant_year is not None:
        for pos, idx in enumerate(tx.index):
            y = int(tx_dates.iloc[pos].year)
            if y != dominant_year:
                wrong_year.append({
                    "Rând în fișier": int(idx) + hdr + 2,
                    "Data din fișier": tx_dates.iloc[pos].date(),
                    "An": y,
                    "Furnizor": str(tx.iloc[pos, furn_idx]).strip() if furn_idx is not None and _has_text(tx.iloc[pos, furn_idx]) else "",
                })
    quality["an_gresit"] = pd.DataFrame(wrong_year)

    # 3) anul din nr. factură diferă de anul rândului (incoerență dată ↔ factură)
    fact_mismatch = []
    if fact_idx is not None:
        for pos, idx in enumerate(tx.index):
            fv = tx.iloc[pos, fact_idx]
            if not _has_text(fv):
                continue
            ys = re.findall(r"\b(20\d{2})\b", str(fv))
            row_year = int(tx_dates.iloc[pos].year)
            if ys and all(int(y) != row_year for y in ys):
                fact_mismatch.append({
                    "Rând în fișier": int(idx) + hdr + 2,
                    "Data rând (an)": row_year,
                    "Nr. factură": str(fv).strip(),
                    "An(i) în factură": ", ".join(sorted(set(ys))),
                })
    quality["factura_an"] = pd.DataFrame(fact_mismatch)

    # 4) HARTA CELULELOR CU PROBLEME (formule lipsă / valori hardcodate ce nu respectă formula)
    cells = []
    ex_rows = np.array([int(i) + hdr + 2 for i in tx.index])

    # 4a) „Total cheltuită" pe rând = suma categoriilor
    tc_letter = _xl_col(cols["total_cheltuit"])
    typed_tc = pd.to_numeric(tx.iloc[:, cols["total_cheltuit"]], errors="coerce").values
    sum_cat = row_cheltuit.values
    for pos in np.where(np.isnan(typed_tc) & (np.abs(sum_cat) > 0.01))[0]:
        cells.append([f"{tc_letter}{ex_rows[pos]}", int(ex_rows[pos]), "Total cheltuită",
                      None, round(float(sum_cat[pos]), 2), None,
                      "Formulă lipsă: celulă goală, dar categoriile însumează o valoare"])
    for pos in np.where((~np.isnan(typed_tc)) & (np.abs(typed_tc - sum_cat) > 0.01))[0]:
        cells.append([f"{tc_letter}{ex_rows[pos]}", int(ex_rows[pos]), "Total cheltuită",
                      round(float(typed_tc[pos]), 2), round(float(sum_cat[pos]), 2),
                      round(float(typed_tc[pos] - sum_cat[pos]), 2),
                      "Valoare hardcodată ≠ suma categoriilor"])

    # 4b) „Sold cont" pe rând = sold precedent + intrări − ieșiri (= soldul curent recalculat)
    if cols["sold"] is not None:
        sd_letter = _xl_col(cols["sold"])
        sold_typed_v = pd.to_numeric(tx.iloc[:, cols["sold"]], errors="coerce").values
        run_v = running.values
        mov_v = movement.values
        anchor_pos = -1
        st_series = pd.to_numeric(tx.iloc[:, cols["sold"]], errors="coerce")
        fvi = st_series.first_valid_index()
        if fvi is not None:
            anchor_pos = int(tx.index.get_loc(fvi))
        for pos in range(len(ex_rows)):
            if pos == anchor_pos:
                continue
            stv, rnv2 = sold_typed_v[pos], run_v[pos]
            if np.isnan(stv):
                if abs(mov_v[pos]) > 0.01:
                    cells.append([f"{sd_letter}{ex_rows[pos]}", int(ex_rows[pos]), "Sold cont",
                                  None, round(float(rnv2), 2), None,
                                  "Formulă lipsă: sold gol, deși rândul are mișcare de bani"])
            elif abs(stv - rnv2) > 0.01:
                cells.append([f"{sd_letter}{ex_rows[pos]}", int(ex_rows[pos]), "Sold cont",
                              round(float(stv), 2), round(float(rnv2), 2),
                              round(float(stv - rnv2), 2),
                              "Sold hardcodat ≠ sold recalculat (sold precedent + intrări − ieșiri)"])

    # 4c) Rândul de TOTAL: fiecare celulă ar trebui să fie suma coloanei peste tranzacții
    if boundary is not None:
        trow = boundary + hdr + 2
        col_expected = {}
        cat_sums = cat_matrix.sum(axis=0)  # indexat pe cat_names
        for k, c in enumerate(cat_idxs):
            col_expected[c] = (float(cat_sums.iloc[k]), cat_names[k])
        col_expected[cols["total_cheltuit"]] = (total_cheltuit, "Total cheltuită")
        if cols["total_primit"] is not None:
            col_expected[cols["total_primit"]] = (float(primit.sum()), "Total primit")
        if cols["venit_clarificare"] is not None:
            col_expected[cols["venit_clarificare"]] = (float(venit.sum()), "Venituri în clarificare")
        if cols["transferat"] is not None:
            col_expected[cols["transferat"]] = (float(transferat.sum()), "Sume transferate")
        for c, (exp, cname) in col_expected.items():
            cellval = pd.to_numeric(pd.Series([body_full.iloc[boundary, c]]), errors="coerce").iloc[0]
            letter = _xl_col(c)
            if pd.isna(cellval):
                if abs(exp) > 0.01:
                    cells.append([f"{letter}{trow}", int(trow), f"TOTAL · {cname}",
                                  None, round(exp, 2), None,
                                  "Formulă lipsă în rândul de total (celulă goală)"])
            elif abs(cellval - exp) > 0.01:
                cells.append([f"{letter}{trow}", int(trow), f"TOTAL · {cname}",
                              round(float(cellval), 2), round(exp, 2), round(float(cellval - exp), 2),
                              "Total hardcodat ≠ suma coloanei"])

    cells_df = pd.DataFrame(cells, columns=[
        "Celulă", "Rând", "Coloană", "Valoare în fișier", "Valoare corectă", "Diferență", "Problemă"])
    quality["celule"] = cells_df
    quality["celule_counts"] = (cells_df["Problemă"].value_counts() if len(cells_df)
                                else pd.Series(dtype=int))

    return {
        "header_row": hdr, "cols": cols, "date_idx": date_idx,
        "raw": raw,
        "tx_detail": tx_detail, "quality": quality,
        "cat_names": cat_names, "n_tx": int(len(tx)),
        "period_min": tx_dates.min(), "period_max": tx_dates.max(),
        "opening": opening, "opening_source": opening_source,
        "total_cheltuit": total_cheltuit,
        "total_primit": float(primit.sum()),
        "total_venit": float(venit.sum()),
        "total_transferat": float(transferat.sum()),
        "final_sold": final_sold,
        "cat_table": cat_table, "month_table": month_table,
        "day_table": day_table,
        "_daily_src": pd.DataFrame({
            "Data": tx_dates.values, "primit": primit.values,
            "venit": venit.values,
            "cheltuit": row_cheltuit.values, "transferat": transferat.values,
        }),
        "supplier_table": supplier_table, "sold_curve": sold_curve,
        "checks": checks, "grand": grand,
    }


def build_export(res):
    """Construiește un Excel curat cu totalurile corecte."""
    buf = io.BytesIO()
    with pd.ExcelWriter(buf, engine="openpyxl") as xl:
        sumar = pd.DataFrame({
            "Indicator": ["Perioadă", "Număr tranzacții", "Sold inițial",
                          "Total primit", "Total cheltuit", "Sume transferate",
                          "Sold final"],
            "Valoare": [
                f"{res['period_min']:%d.%m.%Y} – {res['period_max']:%d.%m.%Y}",
                res["n_tx"], round(res["opening"], 2), round(res["total_primit"], 2),
                round(res["total_cheltuit"], 2), round(res["total_transferat"], 2),
                round(res["final_sold"], 2),
            ],
        })
        sumar.to_excel(xl, sheet_name="Sumar", index=False)
        res["cat_table"].round(2).to_excel(xl, sheet_name="Pe categorii", index=False)
        res["month_table"].round(2).to_excel(xl, sheet_name="Pe luni", index=False)
        if res.get("day_table") is not None:
            res["day_table"].round(2).to_excel(xl, sheet_name="Pe zile", index=False)
        if res["supplier_table"] is not None:
            res["supplier_table"].round(2).to_excel(xl, sheet_name="Pe furnizor", index=False)
    buf.seek(0)
    return buf.getvalue()


def _clean_cell(v):
    """Pregătește o valoare pentru scriere în openpyxl (fără NaN/tipuri numpy)."""
    if v is None:
        return None
    if isinstance(v, float) and np.isnan(v):
        return None
    if isinstance(v, np.integer):
        return int(v)
    if isinstance(v, np.floating):
        return float(v)
    if isinstance(v, (pd.Timestamp, datetime.datetime, datetime.date)):
        try:
            return pd.Timestamp(v).to_pydatetime()
        except Exception:
            return str(v)
    return v


def build_cell_audit_xlsx(res):
    """Reproduce fișierul original și marchează pe hartă celulele cu probleme:
    roșu = valoare hardcodată greșită, galben = formulă lipsă. Fiecare celulă
    marcată are un comentariu cu valoarea corectă."""
    from openpyxl import Workbook
    from openpyxl.styles import PatternFill, Font
    from openpyxl.comments import Comment

    raw = res["raw"]
    cells = res["quality"]["celule"]
    hdr = res["header_row"]

    wb = Workbook()
    ws = wb.active
    ws.title = "Fisier marcat"

    nrows, ncols = raw.shape
    for i in range(nrows):
        rowvals = raw.iloc[i].tolist()
        for j, v in enumerate(rowvals):
            cv = _clean_cell(v)
            if cv is not None:
                ws.cell(row=i + 1, column=j + 1, value=cv)

    # antetul îngroșat + înghețat
    bold = Font(bold=True)
    for j in range(ncols):
        ws.cell(row=hdr + 1, column=j + 1).font = bold
    ws.freeze_panes = ws.cell(row=hdr + 2, column=1)

    red = PatternFill("solid", fgColor="FFC7CE")
    yellow = PatternFill("solid", fgColor="FFEB9C")
    for _, r in cells.iterrows():
        try:
            cell = ws[r["Celulă"]]
        except Exception:
            continue
        is_missing = "lipsă" in str(r["Problemă"]) or "lipsa" in _norm(r["Problemă"])
        cell.fill = yellow if is_missing else red
        corect = r["Valoare corectă"]
        fis = r["Valoare în fișier"]
        msg = (f"{r['Problemă']}\n"
               f"Valoare în fișier: {'(gol)' if pd.isna(fis) else fis}\n"
               f"Valoare corectă: {corect}")
        cell.comment = Comment(msg, "Verificare automată")

    # legendă
    leg = wb.create_sheet("Legendă")
    leg["A1"] = "Roșu = valoare hardcodată greșită (nu respectă formula corectă)"
    leg["A1"].fill = red
    leg["A2"] = "Galben = formulă lipsă (celulă goală care ar fi trebuit calculată)"
    leg["A2"].fill = yellow
    leg["A4"] = "Treci cu mouse-ul peste o celulă marcată ca să vezi valoarea corectă."
    leg.column_dimensions["A"].width = 70

    # lista completă a problemelor
    lst = wb.create_sheet("Lista probleme")
    lst.append(list(cells.columns))
    for _, r in cells.iterrows():
        lst.append([None if (isinstance(x, float) and pd.isna(x)) else x for x in r.tolist()])

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()


# ----------------------------------------------------------------------------- #
#  EXTRAS DE CONT BANCAR (sursa reală de adevăr — independentă de formulele Excel)
# ----------------------------------------------------------------------------- #

def _num(v):
    """Convertește orice celulă la float, tolerant la text și separatori RO/EN."""
    if v is None or (isinstance(v, float) and np.isnan(v)):
        return np.nan
    if isinstance(v, (int, float)):
        return float(v)
    s = str(v).strip().replace("\xa0", "").replace(" ", "")
    if s == "":
        return np.nan
    if "," in s and "." in s:
        # ultimul separator = separatorul zecimal
        if s.rfind(",") > s.rfind("."):
            s = s.replace(".", "").replace(",", ".")
        else:
            s = s.replace(",", "")
    elif "," in s:
        s = s.replace(",", ".")
    try:
        return float(s)
    except ValueError:
        return np.nan


def _bank_date(v):
    """Acceptă atât datetime cât și text 'dd/mm/yyyy' (format extras bancar)."""
    if isinstance(v, (pd.Timestamp, datetime.datetime, datetime.date)):
        try:
            return pd.Timestamp(v)
        except Exception:
            return pd.NaT
    if isinstance(v, str):
        m = re.fullmatch(r"(\d{1,2})[/.-](\d{1,2})[/.-](\d{4})", v.strip())
        if m:
            try:
                return pd.Timestamp(int(m.group(3)), int(m.group(2)), int(m.group(1)))
            except ValueError:
                return pd.NaT
    return pd.NaT


def is_bank_statement(raw):
    """Recunoaște un extras de cont după antetul Data/Debit/Credit + 'sold initial'."""
    head = raw.iloc[:min(60, len(raw))]
    text = " ".join(_norm(v) for v in head.to_numpy().ravel() if v is not None)
    has_extras = "extras cont" in text or "sold initial" in text
    has_cols = "debit" in text and "credit" in text and "descri" in text
    return has_extras and has_cols


def _op_type(descr, debit, credit):
    """Clasifică operațiunea în funcție de descriere și sensul mișcării."""
    s = _norm(descr)
    if credit and credit > 0:
        if "dobanda" in s:
            return "Dobândă încasată"
        return "Încasare (subvenție / retur)"
    if "speze" in s or "comision" in s:
        return "Comisioane și speze bancare"
    return "Plată / transfer"


def _op_benef(descr):
    """Extrage numele beneficiarului dintr-o descriere de tip 'Beneficiar: X Platitor'."""
    m = re.search(r"beneficiar[:\s]+(.*?)\s+platitor", str(descr), flags=re.IGNORECASE)
    if m:
        return re.sub(r"\s+", " ", m.group(1)).strip(" :")
    return None


def parse_bank_statement(raw):
    """Extrage din extras: sold inițial, tranzacțiile (dată/descriere/debit/credit) și
    valorile declarate de bancă (sold final, rulaje) pentru verificare."""
    n = len(raw)

    opening, opening_row = None, None
    stated_final, available = None, None
    for i in range(min(40, n)):
        cells = raw.iloc[i].tolist()
        rn = [_norm(v) for v in cells]
        nums = [x for x in (_num(v) for v in cells) if not pd.isna(x)]
        if any(x == "sold initial" for x in rn) and nums and opening is None:
            opening, opening_row = nums[0], i
        if any("suma disponibila" in x for x in rn) and nums:
            available = nums[0]

    hdr = None
    for i in range(min(80, n)):
        rn = [_norm(v) for v in raw.iloc[i].tolist()]
        if (any(x == "data" for x in rn) and any("debit" in x for x in rn)
                and any("credit" in x for x in rn) and any("descri" in x for x in rn)):
            hdr = i
            break
    if hdr is None:
        raise ValueError("Nu pare extras de cont: nu am găsit antetul Data/Descriere/Debit/Credit.")

    rn = [_norm(v) for v in raw.iloc[hdr].tolist()]

    def col(pred):
        for idx, x in enumerate(rn):
            if pred(x):
                return idx
        return None

    c_data = col(lambda x: x == "data")
    c_desc = col(lambda x: "descri" in x)
    c_deb = col(lambda x: x == "debit" or ("debit" in x and "rulaj" not in x))
    c_cre = col(lambda x: x == "credit" or ("credit" in x and "rulaj" not in x))
    if None in (c_data, c_deb, c_cre):
        raise ValueError("Extrasul nu are coloanele așteptate (Data / Debit / Credit).")

    recs = []
    rulaj_deb = rulaj_cre = None
    for i in range(hdr + 1, n):
        cells = raw.iloc[i].tolist()
        rowtext = " ".join(_norm(v) for v in cells)
        if "sold final" in rowtext:
            nums = [x for x in (_num(v) for v in cells) if not pd.isna(x)]
            if nums:
                stated_final = nums[0]
            break  # tranzacțiile se opresc aici; mai jos urmează sinteza zilnică
        if "rulaj" in rowtext:
            nums = [x for x in (_num(v) for v in cells) if not pd.isna(x)]
            if len(nums) >= 2:
                rulaj_deb, rulaj_cre = nums[0], nums[1]
            continue
        d = _bank_date(raw.iloc[i, c_data])
        if pd.isna(d):
            continue
        deb = _num(raw.iloc[i, c_deb])
        cre = _num(raw.iloc[i, c_cre])
        recs.append((d,
                     "" if c_desc is None else ("" if pd.isna(raw.iloc[i, c_desc]) else str(raw.iloc[i, c_desc])),
                     0.0 if pd.isna(deb) else deb,
                     0.0 if pd.isna(cre) else cre))

    if not recs:
        raise ValueError("Nu am găsit nicio tranzacție în extras.")

    tx = pd.DataFrame(recs, columns=["Data", "Descriere", "Debit", "Credit"])
    return {
        "opening": opening, "opening_row": opening_row,
        "tx": tx, "stated_final": stated_final, "available": available,
        "rulaj_deb": rulaj_deb, "rulaj_cre": rulaj_cre,
    }


def analyze_bank(raw):
    """Analiză pas-cu-pas a extrasului. Recalculează TOTUL din debit/credit,
    fără să se bazeze pe vreo formulă din fișier; verifică apoi față de ce declară banca."""
    p = parse_bank_statement(raw)
    tx = p["tx"].sort_values("Data").reset_index(drop=True)
    opening = float(p["opening"]) if p["opening"] is not None else 0.0
    opening_source = "preluat din extras (SOLD INIȚIAL)" if p["opening"] is not None else "presupus 0 (nu am găsit SOLD INIȚIAL)"

    tx["Net"] = tx["Credit"] - tx["Debit"]
    tx["Sold"] = opening + tx["Net"].cumsum()
    tx["Tip"] = [_op_type(d, db, cr) for d, db, cr in zip(tx["Descriere"], tx["Debit"], tx["Credit"])]
    tx["Beneficiar"] = tx["Descriere"].map(_op_benef)

    total_debit = float(tx["Debit"].sum())
    total_credit = float(tx["Credit"].sum())
    final_calc = opening + total_credit - total_debit

    # ---- pe luni ----
    per = tx["Data"].dt.to_period("M")
    mdf = pd.DataFrame({"p": per.values, "Debit": tx["Debit"].values,
                        "Credit": tx["Credit"].values, "Net": tx["Net"].values})
    g = mdf.groupby("p", sort=True).sum(numeric_only=True)
    g["sold_sf"] = opening + g["Net"].cumsum()
    month_table = pd.DataFrame({
        "Luna": [_luna_ro(pp) for pp in g.index],
        "Intrări (lei)": g["Credit"].values,
        "Ieșiri (lei)": g["Debit"].values,
        "Net (lei)": g["Net"].values,
        "Sold la sfârșit de lună (lei)": g["sold_sf"].values,
    })

    # ---- pe zile ----
    ddf = pd.DataFrame({"zi": tx["Data"].dt.normalize().values, "Debit": tx["Debit"].values,
                        "Credit": tx["Credit"].values, "Net": tx["Net"].values})
    gd = ddf.groupby("zi", sort=True).sum(numeric_only=True)
    gd["sold_sf"] = opening + gd["Net"].cumsum()
    day_table = pd.DataFrame({
        "Data": [pd.Timestamp(z).date() for z in gd.index],
        "Intrări (lei)": gd["Credit"].values,
        "Ieșiri (lei)": gd["Debit"].values,
        "Net (lei)": gd["Net"].values,
        "Sold la sfârșit de zi (lei)": gd["sold_sf"].values,
    })

    # ---- pe tip de operațiune ----
    type_out = (tx[tx["Debit"] > 0].groupby("Tip", as_index=False)["Debit"].sum()
                .sort_values("Debit", ascending=False)
                .rename(columns={"Debit": "Total ieșit (lei)"}))
    type_in = (tx[tx["Credit"] > 0].groupby("Tip", as_index=False)["Credit"].sum()
               .sort_values("Credit", ascending=False)
               .rename(columns={"Credit": "Total intrat (lei)"}))

    # ---- pe beneficiar (doar plăți / ieșiri) ----
    pay = tx[(tx["Debit"] > 0) & tx["Beneficiar"].notna()].copy()
    benef_table = (pay.groupby("Beneficiar", as_index=False)["Debit"].sum()
                   .sort_values("Debit", ascending=False)
                   .rename(columns={"Debit": "Total plătit (lei)"})) if len(pay) else None

    fees = float(tx.loc[tx["Tip"] == "Comisioane și speze bancare", "Debit"].sum())
    interest = float(tx.loc[tx["Tip"] == "Dobândă încasată", "Credit"].sum())

    # ---- verificări față de ce declară banca ----
    verify = {}
    if p["stated_final"] is not None:
        verify["sold_final"] = {"calculat": final_calc, "din_extras": float(p["stated_final"]),
                                "diferenta": final_calc - float(p["stated_final"])}
    if p["available"] is not None:
        verify["disponibil"] = {"calculat": final_calc, "din_extras": float(p["available"]),
                                "diferenta": final_calc - float(p["available"])}
    if p["rulaj_deb"] is not None:
        verify["rulaj_debit"] = {"calculat": total_debit, "din_extras": float(p["rulaj_deb"]),
                                 "diferenta": total_debit - float(p["rulaj_deb"])}
    if p["rulaj_cre"] is not None:
        verify["rulaj_credit"] = {"calculat": total_credit, "din_extras": float(p["rulaj_cre"]),
                                  "diferenta": total_credit - float(p["rulaj_cre"])}

    sold_curve = tx[["Data", "Sold"]].rename(columns={"Sold": "Sold (lei)"}).dropna()

    return {
        "kind": "extras",
        "opening": opening, "opening_source": opening_source,
        "n_tx": int(len(tx)),
        "period_min": tx["Data"].min(), "period_max": tx["Data"].max(),
        "total_credit": total_credit, "total_debit": total_debit,
        "final_calc": final_calc, "fees": fees, "interest": interest,
        "month_table": month_table, "day_table": day_table,
        "type_out": type_out, "type_in": type_in,
        "benef_table": benef_table, "sold_curve": sold_curve,
        "verify": verify, "tx": tx,
    }


def build_export_bank(res):
    """Excel curat cu rezultatele recalculate din extras."""
    buf = io.BytesIO()
    with pd.ExcelWriter(buf, engine="openpyxl") as xl:
        sumar = pd.DataFrame({
            "Indicator": ["Perioadă", "Număr operațiuni", "Sold inițial",
                          "Total intrări (credit)", "Total ieșiri (debit)",
                          "din care comisioane/speze", "din care dobândă încasată",
                          "Sold final calculat"],
            "Valoare": [
                f"{res['period_min']:%d.%m.%Y} – {res['period_max']:%d.%m.%Y}",
                res["n_tx"], round(res["opening"], 2), round(res["total_credit"], 2),
                round(res["total_debit"], 2), round(res["fees"], 2),
                round(res["interest"], 2), round(res["final_calc"], 2),
            ],
        })
        sumar.to_excel(xl, sheet_name="Sumar extras", index=False)
        res["month_table"].round(2).to_excel(xl, sheet_name="Pe luni", index=False)
        if res.get("day_table") is not None:
            res["day_table"].round(2).to_excel(xl, sheet_name="Pe zile", index=False)
        res["type_out"].round(2).to_excel(xl, sheet_name="Ieșiri pe tip", index=False)
        res["type_in"].round(2).to_excel(xl, sheet_name="Intrări pe tip", index=False)
        if res["benef_table"] is not None:
            res["benef_table"].round(2).to_excel(xl, sheet_name="Pe beneficiar", index=False)
    buf.seek(0)
    return buf.getvalue()


def reconcile(report, bank):
    """Compară cele două surse independente la punctele care TREBUIE să coincidă."""
    rows = []

    rows.append({
        "Indicator": "Sold inițial",
        "Din raport": report["opening"],
        "Din extras (bancă)": bank["opening"],
        "Diferență": report["opening"] - bank["opening"],
    })

    bank_final = bank["final_calc"]
    rows.append({
        "Indicator": "Sold final (recalculat)",
        "Din raport": report["final_sold"],
        "Din extras (bancă)": bank_final,
        "Diferență": report["final_sold"] - bank_final,
    })

    # ieșiri totale din bancă vs (cheltuieli + transferuri + comisioane) din raport
    report_out = report["total_cheltuit"] + report["total_transferat"]
    rows.append({
        "Indicator": "Ieșiri (cheltuieli + transferuri vs debite bancă)",
        "Din raport": report_out,
        "Din extras (bancă)": bank["total_debit"],
        "Diferență": report_out - bank["total_debit"],
    })
    rows.append({
        "Indicator": "Intrări (primit + venituri vs credite bancă)",
        "Din raport": report["total_primit"] + report["total_venit"],
        "Din extras (bancă)": bank["total_credit"],
        "Diferență": (report["total_primit"] + report["total_venit"]) - bank["total_credit"],
    })

    return pd.DataFrame(rows)


_NAME_SUFFIXES = {"srl", "sa", "sas", "pfa", "srld", "srls", "snc", "scs", "sc", "ii", "srl-d"}


def _name_key(name):
    """Set de cuvinte-cheie din numele furnizorului (fără diacritice, fără SRL/SA etc.)."""
    s = _norm(name)
    s = re.sub(r"[^a-z0-9 ]", " ", s)
    return {w for w in s.split() if w and w not in _NAME_SUFFIXES and len(w) > 1}


def _name_sim(a, b):
    if not a or not b:
        return 0.0
    return len(a & b) / len(a | b)


def reconcile_payments(report, bank, tol=0.01):
    """Potrivire plată-cu-plată între raport și extras, pe (sumă + furnizor).
    Întoarce: potriviri, plăți din extras fără corespondent în raport,
    rânduri din raport fără corespondent în extras."""
    rep = report["tx_detail"].copy()
    rep = rep[rep["Furnizor"].map(lambda v: bool(str(v).strip()))]
    rep = rep[pd.to_numeric(rep["Sumă (lei)"], errors="coerce").fillna(0) != 0].reset_index(drop=True)

    sd_letter = _xl_col(report["cols"]["sold"]) if report["cols"].get("sold") is not None else None

    tx = bank["tx"]
    bank_pay = tx[(tx["Debit"] > 0) & tx["Beneficiar"].notna()].copy().reset_index(drop=True)

    rep_keys = [_name_key(n) for n in rep["Furnizor"]]
    rep_amt = [round(float(a), 2) for a in rep["Sumă (lei)"]]
    bk_keys = [_name_key(n) for n in bank_pay["Beneficiar"]]
    bk_amt = [round(float(a), 2) for a in bank_pay["Debit"]]

    from collections import defaultdict
    by_amt = defaultdict(list)
    for j, a in enumerate(bk_amt):
        by_amt[a].append(j)

    used = set()
    matches = []
    rep_unmatched_idx = []

    def _try_match(strict):
        for i in range(len(rep)):
            if i in matched_rep:
                continue
            cands = [j for j in by_amt.get(rep_amt[i], []) if j not in used]
            if not cands:
                continue
            best, bs = None, -1.0
            for j in cands:
                s = _name_sim(rep_keys[i], bk_keys[j])
                if s > bs:
                    bs, best = s, j
            ok = (bs >= 0.34 or rep_keys[i] <= bk_keys[best] or bk_keys[best] <= rep_keys[i]) if strict else True
            if best is not None and ok:
                used.add(best)
                matched_rep.add(i)
                rep_sold = rep.iloc[i].get("Sold AG (raport)")
                bank_sold = bank_pay.iloc[best]["Sold"]
                matches.append({
                    "Furnizor (raport)": rep.iloc[i]["Furnizor"],
                    "Beneficiar (bancă)": bank_pay.iloc[best]["Beneficiar"],
                    "Sumă (lei)": rep_amt[i],
                    "Nr. factură (raport)": rep.iloc[i]["Nr. factură"],
                    "Data (bancă)": pd.Timestamp(bank_pay.iloc[best]["Data"]).date(),
                    "Rând raport": int(rep.iloc[i]["Rând"]),
                    "Sold AG (raport)": (round(float(rep_sold), 2) if pd.notna(rep_sold) else None),
                    "Sold real (bancă)": round(float(bank_sold), 2),
                    "Diferență sold": (round(float(rep_sold) - float(bank_sold), 2) if pd.notna(rep_sold) else None),
                    "Celulă AG": (f"{sd_letter}{int(rep.iloc[i]['Rând'])}" if sd_letter else None),
                    "Potrivire nume": round(bs, 2) if strict else 0.0,
                })

    matched_rep = set()
    _try_match(strict=True)
    _try_match(strict=False)  # al doilea pas: aceeași sumă, nume diferit

    for i in range(len(rep)):
        if i not in matched_rep:
            rep_unmatched_idx.append(i)

    rep_unmatched = pd.DataFrame([{
        "Rând în fișier": int(rep.iloc[i]["Rând"]),
        "Data": pd.Timestamp(rep.iloc[i]["Data"]).date() if pd.notna(rep.iloc[i]["Data"]) else None,
        "Furnizor": rep.iloc[i]["Furnizor"],
        "Nr. factură": rep.iloc[i]["Nr. factură"],
        "Sumă (lei)": rep_amt[i],
    } for i in rep_unmatched_idx])

    bank_unmatched = pd.DataFrame([{
        "Data": pd.Timestamp(bank_pay.iloc[j]["Data"]).date(),
        "Beneficiar": bank_pay.iloc[j]["Beneficiar"],
        "Sumă (lei)": bk_amt[j],
        "Descriere": str(bank_pay.iloc[j]["Descriere"])[:120],
    } for j in range(len(bank_pay)) if j not in used])

    matches_df = pd.DataFrame(matches)

    # Verificarea soldului AG (raport) față de soldul real din bancă, la plățile potrivite
    sold_check = None
    sold_ok = sold_bad = 0
    if len(matches_df) and "Diferență sold" in matches_df.columns:
        have = matches_df[matches_df["Diferență sold"].notna()].copy()
        if len(have):
            have = have.sort_values("Data (bancă)")
            sold_ok = int((have["Diferență sold"].abs() <= 0.01).sum())
            sold_bad = int((have["Diferență sold"].abs() > 0.01).sum())
            sold_check = have[["Data (bancă)", "Furnizor (raport)", "Sumă (lei)", "Celulă AG",
                               "Sold AG (raport)", "Sold real (bancă)", "Diferență sold"]]

    return {
        "matches": matches_df,
        "rep_unmatched": rep_unmatched,
        "bank_unmatched": bank_unmatched,
        "n_rep": len(rep), "n_bank": len(bank_pay), "n_match": len(matches_df),
        "sold_check": sold_check, "sold_ok": sold_ok, "sold_bad": sold_bad,
    }


def daily_diff(report, bank, tol=0.01):
    """Raport de diferențe PE ZI între raport și extras.
    Aliniem pe ziua calendaristică (extrasul e pe un an; în raport ziua/luna sunt
    corecte chiar dacă anul e greșit, așa că forțăm anul raportului la anul extrasului).
    Ieșirile includ de ambele părți comisioanele (în raport ca total zilnic pe categoria
    «com. bancare», în extras ca linii de speze)."""
    bank_year = int(pd.Timestamp(bank["period_max"]).year)

    src = report["_daily_src"].copy()
    src["Data"] = pd.to_datetime(src["Data"])

    def _to_bank_year(d):
        try:
            return pd.Timestamp(year=bank_year, month=d.month, day=d.day)
        except ValueError:
            return pd.Timestamp(year=bank_year, month=d.month, day=28)

    src["zi"] = src["Data"].map(_to_bank_year)
    src["intrari"] = src["primit"] + (src["venit"] if "venit" in src.columns else 0.0)
    src["iesiri"] = src["cheltuit"] + src["transferat"]
    rep_day = src.groupby("zi").agg(
        Primit_raport=("intrari", "sum"),
        Iesiri_raport=("iesiri", "sum"),
    )

    bk = bank["tx"].copy()
    bk["zi"] = pd.to_datetime(bk["Data"]).dt.normalize()
    bank_day = bk.groupby("zi").agg(
        Intrari_banca=("Credit", "sum"),
        Iesiri_banca=("Debit", "sum"),
    )

    alld = rep_day.join(bank_day, how="outer").fillna(0.0).sort_index()

    # sold la sfârșitul fiecărei zile (recalculat de ambele părți, în ordine calendaristică)
    opening_rep = float(report["opening"])
    opening_bank = float(bank["opening"])
    rep_mov = alld["Primit_raport"] - alld["Iesiri_raport"]
    bank_mov = alld["Intrari_banca"] - alld["Iesiri_banca"]
    sold_rep = opening_rep + rep_mov.cumsum()
    sold_bank = opening_bank + bank_mov.cumsum()

    out = pd.DataFrame({
        "Data": [pd.Timestamp(z).date() for z in alld.index],
        "Intrări raport (lei)": alld["Primit_raport"].values,
        "Intrări bancă (lei)": alld["Intrari_banca"].values,
        "Dif. intrări": (alld["Primit_raport"] - alld["Intrari_banca"]).values,
        "Ieșiri raport (lei)": alld["Iesiri_raport"].values,
        "Ieșiri bancă (lei)": alld["Iesiri_banca"].values,
        "Dif. ieșiri": (alld["Iesiri_raport"] - alld["Iesiri_banca"]).values,
        "Sold raport (lei)": sold_rep.values,
        "Sold bancă (lei)": sold_bank.values,
        "Dif. sold": (sold_rep - sold_bank).values,
    })
    # „diferențe" = la flux (intrări/ieșiri); soldul e cumulativ, îl raportăm separat
    bad = out[(out["Dif. intrări"].abs() > tol) | (out["Dif. ieșiri"].abs() > tol)]
    return {"full": out, "bad": bad,
            "n_zile": len(out), "n_zile_dif": len(bad),
            "max_dif_sold": float(out["Dif. sold"].abs().max()) if len(out) else 0.0,
            "final_dif_sold": float(out["Dif. sold"].iloc[-1]) if len(out) else 0.0,
            "bank_year": bank_year}


# ----------------------------------------------------------------------------- #
#  INTERFAȚĂ STREAMLIT
# ----------------------------------------------------------------------------- #

def _fmt(x):
    return f"{x:,.2f}".replace(",", " ").replace(".", ",") + " lei"


def render_report(st, res):
    # ---- 1. SUMAR ----
    st.header("1. Sumar general")
    st.caption(f"Perioadă acoperită: **{res['period_min']:%d.%m.%Y} – {res['period_max']:%d.%m.%Y}** · "
               f"**{res['n_tx']}** tranzacții · {len(res['cat_names'])} categorii de cheltuieli detectate.")
    c1, c2, c3, c4 = st.columns(4)
    c1.metric("Total primit", _fmt(res["total_primit"]))
    c2.metric("Total cheltuit", _fmt(res["total_cheltuit"]))
    c3.metric("Sume transferate", _fmt(res["total_transferat"]))
    c4.metric("Sold final", _fmt(res["final_sold"]))
    st.caption(f"Sold inițial folosit: **{_fmt(res['opening'])}** ({res['opening_source']}).")

    st.divider()

    # ---- 2. PE CATEGORII ----
    st.header("2. Cheltuieli pe categorii")
    st.caption("Cât s-a cheltuit pe fiecare categorie, pe toată perioada. Sortat descrescător.")
    ct = res["cat_table"]
    st.dataframe(
        ct.style.format({"Total (lei)": "{:,.2f}", "% din cheltuieli": "{:.1f}%"}),
        use_container_width=True, hide_index=True,
    )
    st.bar_chart(ct.set_index("Categorie")["Total (lei)"], horizontal=True)

    st.divider()

    # ---- 3. PE LUNI ----
    st.header("3. Pe luni")
    st.caption("Pentru fiecare lună: cât s-a primit, cât s-a cheltuit și soldul la sfârșitul lunii.")
    mt = res["month_table"]
    st.dataframe(
        mt.style.format({c: "{:,.2f}" for c in mt.columns if c != "Luna"}),
        use_container_width=True, hide_index=True,
    )

    # ---- pe zile ----
    dt = res.get("day_table")
    if dt is not None and len(dt):
        st.subheader("Pe zile")
        st.caption("Pentru fiecare zi cu mișcări: cât s-a primit, cât s-a cheltuit și soldul la sfârșitul zilei.")
        st.dataframe(
            dt.style.format({c: "{:,.2f}" for c in dt.columns if c != "Data"}),
            use_container_width=True, hide_index=True,
        )
        st.download_button("⬇️ Descarcă soldurile pe zile (CSV)",
                           data=dt.to_csv(index=False).encode("utf-8-sig"),
                           file_name="solduri_pe_zile_raport.csv", mime="text/csv", key="dl_zile_rep")

    st.divider()

    # ---- 4. EVOLUȚIA SOLDULUI ----
    st.header("4. Evoluția soldului în timp")
    st.caption("Soldul contului după fiecare tranzacție (recalculat corect).")
    st.line_chart(res["sold_curve"].set_index("Data")["Sold (lei)"])

    st.divider()

    # ---- 5. PE FURNIZOR ----
    if res["supplier_table"] is not None and len(res["supplier_table"]):
        st.header("5. Pe furnizor")
        st.caption("Top furnizori după suma totală cheltuită.")
        top = st.slider("Câți furnizori să afișez?", 5, 50, 15)
        sup = res["supplier_table"].head(top)
        st.dataframe(
            sup.style.format({"Total cheltuit (lei)": "{:,.2f}"}),
            use_container_width=True, hide_index=True,
        )
        st.divider()

    # ---- 6. VERIFICĂRI ----
    st.header("6. Verificări — unde fișierul ar putea avea probleme")
    st.caption("Acestea sunt diferențe între ce e scris în fișier și calculul corect. "
               "Nu modifică fișierul — doar îți arată ce merită verificat.")
    checks = res["checks"]

    # headline dublă-numărare
    if res.get("grand"):
        g = res["grand"]
        diff = g["total_fisier"] - res["total_cheltuit"]
        if abs(diff) > 0.01:
            st.warning(
                f"**Totalul general din fișier (rândul {g['rand']}) = {_fmt(g['total_fisier'])}**, "
                f"dar totalul corect recalculat = **{_fmt(res['total_cheltuit'])}**. "
                f"Diferență: {_fmt(diff)}. (De obicei înseamnă dublă-numărare în rândul de total.)"
            )
        else:
            st.success(f"Totalul general din fișier coincide cu calculul corect ({_fmt(res['total_cheltuit'])}).")

    def show_check(key, ok_msg, bad_title):
        df = checks.get(key)
        if df is None:
            return
        if len(df) == 0:
            st.success(ok_msg)
        else:
            with st.expander(f"⚠️ {bad_title} — {len(df)} rând(uri)"):
                st.dataframe(df, use_container_width=True, hide_index=True)

    show_check("total_rand_gresit",
               "Pe fiecare rând, totalul tastat coincide cu suma categoriilor.",
               "Rânduri unde totalul tastat diferă de suma categoriilor")
    show_check("orfane",
               "Toate cheltuielile au dată — nimic exclus din calcul.",
               "Tranzacții cu cheltuieli dar FĂRĂ dată (NU au fost numărate — adaugă data în fișier)")

    sold = checks.get("sold")
    if sold is not None:
        if sold["n"] == 0:
            st.success("Soldul din fișier coincide cu soldul recalculat pe tot parcursul.")
        else:
            st.warning(
                f"**Soldul din fișier diferă de cel corect.** Prima diferență apare la "
                f"**rândul {sold['first_row']}** (fișier: {_fmt(sold['first_typed'])}, "
                f"corect: {_fmt(sold['first_correct'])}). La final, fișierul arată "
                f"**{_fmt(sold['final_typed'])}**, iar soldul corect este "
                f"**{_fmt(sold['final_correct'])}**."
            )

    st.divider()

    # ---- 7. CALITATEA DATELOR ----
    render_quality(st, res)

    st.divider()

    # ---- 8. EXPORT ----
    st.header("8. Descarcă rezumatul")
    st.caption("Un fișier Excel curat, cu totalurile corecte (sumar, pe categorii, pe luni, pe furnizor).")
    st.download_button(
        "⬇️ Descarcă rezumat Excel",
        data=build_export(res),
        file_name="rezumat_subventie.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        key="dl_report",
    )


def render_quality(st, res):
    q = res.get("quality", {})
    st.header("7. Calitatea datelor — celule lipsă și incoerențe")
    st.caption("Rânduri/celule care nu au valori sau care se contrazic. Util pentru curățarea fișierului.")

    # ---- HARTA CELULELOR CU PROBLEME ----
    cells = q.get("celule")
    if cells is not None:
        st.subheader("Harta celulelor cu probleme (formule lipsă / valori hardcodate)")
        st.caption("Fiecare celulă derivată (Total cheltuită, Sold cont, rândul de Total) este recalculată "
                   "după formula corectă și comparată cu ce e scris în fișier. Mai jos vezi adresa exactă "
                   "a fiecărei celule greșite, valoarea din fișier și valoarea corectă.")
        if len(cells) == 0:
            st.success("Nicio celulă derivată nu are probleme — totul respectă formulele.")
        else:
            counts = q.get("celule_counts")
            st.error(f"**{len(cells)} celule** nu respectă formula corectă.")
            if counts is not None and len(counts):
                st.markdown("**Pe tip de problemă:**")
                st.dataframe(counts.rename_axis("Problemă").reset_index(name="Nr. celule"),
                             use_container_width=True, hide_index=True)
            preview = cells.head(500)
            st.caption(f"Primele {len(preview)} (din {len(cells)}). Descarcă lista completă mai jos.")
            st.dataframe(
                preview.style.format({c: "{:,.2f}" for c in ["Valoare în fișier", "Valoare corectă", "Diferență"]},
                                     na_rep="(gol)"),
                use_container_width=True, hide_index=True,
            )
            st.download_button("⬇️ Descarcă TOATE celulele cu probleme (CSV)",
                               data=cells.to_csv(index=False).encode("utf-8-sig"),
                               file_name="celule_cu_probleme.csv", mime="text/csv", key="dl_cells")
            st.markdown("**Sau descarcă fișierul original cu celulele marcate pe hartă** "
                        "(🔴 roșu = hardcodat greșit, 🟡 galben = formulă lipsă; valoarea corectă apare în comentariul celulei):")
            with st.spinner("Construiesc fișierul marcat…"):
                audit_xlsx = build_cell_audit_xlsx(res)
            st.download_button("⬇️ Descarcă fișierul cu celulele marcate (Excel)",
                               data=audit_xlsx,
                               file_name="fisier_cu_celule_marcate.xlsx",
                               mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                               key="dl_audit")

    # ani
    yc = q.get("year_counts")
    dom = q.get("dominant_year")
    an = q.get("an_gresit")
    if yc is not None and len(yc):
        dist = " · ".join(f"{int(y)}: {int(c)}" for y, c in yc.items())
        st.markdown(f"**Distribuția anilor din coloana de dată:** {dist}.")
        if an is not None and len(an):
            st.warning(
                f"Anul dominant este **{dom}**, dar **{len(an)} rânduri** au altă dată. "
                f"Dacă tot fișierul ar trebui să fie pe {dom}, acestea sunt date greșite (probabil anul tastat greșit)."
            )
            with st.expander(f"Vezi cele {len(an)} rânduri cu an diferit de {dom}"):
                st.dataframe(an, use_container_width=True, hide_index=True)
            st.download_button("⬇️ Descarcă rândurile cu an greșit (CSV)",
                               data=an.to_csv(index=False).encode("utf-8-sig"),
                               file_name="randuri_an_gresit.csv", mime="text/csv", key="dl_an")
        else:
            st.success(f"Toate datele sunt pe anul {dom}.")

    # factură vs an rând
    fa = q.get("factura_an")
    if fa is not None:
        if len(fa):
            st.warning(f"**{len(fa)} rânduri** au în numărul facturii un an diferit de anul datei rândului "
                       "(posibil dată greșită pe rând).")
            with st.expander(f"Vezi cele {len(fa)} incoerențe dată ↔ factură"):
                st.dataframe(fa, use_container_width=True, hide_index=True)
        else:
            st.success("Anul din numărul facturii coincide cu anul datei rândului peste tot.")

    # celule lipsă
    lip = q.get("lipsuri")
    if lip is not None:
        if len(lip):
            st.warning(f"**{len(lip)} rânduri** au cel puțin un câmp lipsă (dată / furnizor / nr. factură / sumă).")
            with st.expander(f"Vezi cele {len(lip)} rânduri cu câmpuri lipsă"):
                st.dataframe(lip, use_container_width=True, hide_index=True)
            st.download_button("⬇️ Descarcă rândurile cu câmpuri lipsă (CSV)",
                               data=lip.to_csv(index=False).encode("utf-8-sig"),
                               file_name="randuri_cu_lipsuri.csv", mime="text/csv", key="dl_lip")
        else:
            st.success("Nu există câmpuri lipsă în rândurile de tranzacții.")


def render_bank(st, res):
    st.header("1. Sumar general (recalculat din extras)")
    st.caption(f"Perioadă: **{res['period_min']:%d.%m.%Y} – {res['period_max']:%d.%m.%Y}** · "
               f"**{res['n_tx']}** operațiuni bancare.")
    c1, c2, c3, c4 = st.columns(4)
    c1.metric("Sold inițial", _fmt(res["opening"]))
    c2.metric("Total intrări", _fmt(res["total_credit"]))
    c3.metric("Total ieșiri", _fmt(res["total_debit"]))
    c4.metric("Sold final calculat", _fmt(res["final_calc"]))

    st.divider()

    # ---- 2. CALCULUL PAS CU PAS ----
    st.header("2. Cum a ieșit soldul final — pas cu pas")
    st.caption("Nu folosim nicio formulă din fișier. Pornim de la soldul inițial și aplicăm, "
               "rând cu rând, intrările (credit) și ieșirile (debit) din extras.")
    steps = pd.DataFrame({
        "Pas": [
            "1. Sold inițial (din extras)",
            "2. + Total intrări (toate creditele)",
            "3. − Total ieșiri (toate debitele)",
            "4. = Sold final calculat de noi",
        ],
        "Sumă (lei)": [res["opening"], res["total_credit"], -res["total_debit"], res["final_calc"]],
    })
    st.dataframe(steps.style.format({"Sumă (lei)": "{:,.2f}"}),
                 use_container_width=True, hide_index=True)
    st.caption(f"Sold inițial: {res['opening_source']}. "
               f"Din ieșiri, **{_fmt(res['fees'])}** sunt comisioane/speze bancare; "
               f"din intrări, **{_fmt(res['interest'])}** sunt dobândă încasată.")

    # verificare față de ce declară banca
    v = res["verify"]
    if v:
        st.subheader("Verificare față de ce declară banca")
        ok = True
        for label, key in [("Sold final", "sold_final"), ("Sumă disponibilă", "disponibil"),
                           ("Rulaj debit (total ieșiri)", "rulaj_debit"),
                           ("Rulaj credit (total intrări)", "rulaj_credit")]:
            d = v.get(key)
            if not d:
                continue
            if abs(d["diferenta"]) < 0.01:
                st.success(f"{label}: calculul nostru = {_fmt(d['calculat'])} coincide cu extrasul.")
            else:
                ok = False
                st.warning(f"{label}: calculat {_fmt(d['calculat'])} vs extras {_fmt(d['din_extras'])} "
                           f"(diferență {_fmt(d['diferenta'])}).")
        if ok:
            st.info("Extrasul este coerent intern — soldul declarat se obține exact din mișcările bancare.")

    st.divider()

    # ---- 3. PE LUNI ----
    st.header("3. Pe luni")
    st.caption("Intrări, ieșiri și soldul la sfârșitul fiecărei luni — toate recalculate.")
    mt = res["month_table"]
    st.dataframe(mt.style.format({c: "{:,.2f}" for c in mt.columns if c != "Luna"}),
                 use_container_width=True, hide_index=True)

    # ---- pe zile ----
    dt = res.get("day_table")
    if dt is not None and len(dt):
        st.subheader("Pe zile")
        st.caption("Intrări, ieșiri și soldul la sfârșitul fiecărei zile cu mișcări — recalculate din extras.")
        st.dataframe(dt.style.format({c: "{:,.2f}" for c in dt.columns if c != "Data"}),
                     use_container_width=True, hide_index=True)
        st.download_button("⬇️ Descarcă soldurile pe zile (CSV)",
                           data=dt.to_csv(index=False).encode("utf-8-sig"),
                           file_name="solduri_pe_zile_extras.csv", mime="text/csv", key="dl_zile_bank")

    st.divider()

    # ---- 4. EVOLUȚIA SOLDULUI ----
    st.header("4. Evoluția soldului în timp")
    st.caption("Soldul contului după fiecare operațiune (calculat de noi din debit/credit).")
    st.line_chart(res["sold_curve"].set_index("Data")["Sold (lei)"])

    st.divider()

    # ---- 5. PE TIP DE OPERAȚIUNE ----
    st.header("5. Pe tip de operațiune")
    cc1, cc2 = st.columns(2)
    with cc1:
        st.caption("Ieșiri (debite) pe tip")
        st.dataframe(res["type_out"].style.format({"Total ieșit (lei)": "{:,.2f}"}),
                     use_container_width=True, hide_index=True)
    with cc2:
        st.caption("Intrări (credite) pe tip")
        st.dataframe(res["type_in"].style.format({"Total intrat (lei)": "{:,.2f}"}),
                     use_container_width=True, hide_index=True)

    st.divider()

    # ---- 6. PE BENEFICIAR ----
    if res["benef_table"] is not None and len(res["benef_table"]):
        st.header("6. Pe beneficiar (cui s-au făcut plățile)")
        st.caption("Top beneficiari după suma totală plătită, extrași automat din descrierea fiecărei plăți.")
        top = st.slider("Câți beneficiari să afișez?", 5, 50, 15, key="bank_topn")
        st.dataframe(res["benef_table"].head(top).style.format({"Total plătit (lei)": "{:,.2f}"}),
                     use_container_width=True, hide_index=True)
        st.divider()

    # ---- 7. EXPORT ----
    st.header("7. Descarcă rezumatul")
    st.download_button(
        "⬇️ Descarcă rezumat extras (Excel)",
        data=build_export_bank(res),
        file_name="rezumat_extras.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        key="dl_bank",
    )


def render_reconcile(st, report, bank):
    st.header("Reconciliere raport ↔ extras de cont")
    st.caption("Comparăm cele două calcule complet independente. Extrasul de cont este sursa reală "
               "(banii care chiar au intrat/ieșit din cont); raportul ar trebui să coincidă cu el.")

    if report["period_min"].date() != bank["period_min"].date() or report["period_max"].date() != bank["period_max"].date():
        st.info(f"Atenție: perioadele diferă — raport: "
                f"{report['period_min']:%d.%m.%Y}–{report['period_max']:%d.%m.%Y}, "
                f"extras: {bank['period_min']:%d.%m.%Y}–{bank['period_max']:%d.%m.%Y}. "
                "Punctele care trebuie oricum să coincidă sunt soldul inițial și soldul final (aceeași dată de sfârșit).")

    rec = reconcile(report, bank)
    st.dataframe(
        rec.style.format({c: "{:,.2f}" for c in rec.columns if c != "Indicator"})
           .map(lambda x: "color: #c00; font-weight: 600" if isinstance(x, (int, float)) and abs(x) > 0.01 else "",
                subset=["Diferență"]),
        use_container_width=True, hide_index=True,
    )

    sold_diff = report["final_sold"] - bank["final_calc"]
    if abs(sold_diff) > 0.01:
        st.warning(
            f"**Soldul final recalculat din raport ({_fmt(report['final_sold'])}) NU coincide cu "
            f"soldul real din bancă ({_fmt(bank['final_calc'])}).** Diferență: {_fmt(sold_diff)}. "
            "Asta confirmă că totalurile din raport (categoriile/formulele) nu reflectă exact mișcările reale de cont — "
            "extrasul este cel corect."
        )
    else:
        st.success("Soldul final din raport coincide cu cel din bancă — raportul este corect la nivel de sold.")

    st.divider()

    # ---- Potrivire plată-cu-plată ----
    st.subheader("Potrivire plată-cu-plată (factură ↔ extras)")
    st.caption("Potrivim fiecare plată din raport cu o plată reală din extras, pe **sumă + furnizor** "
               "(numerele de factură din raport și din extras nu au același format, deci suma + furnizorul "
               "identifică plata cel mai sigur). Mai jos vezi exact ce nu se potrivește.")

    rp = reconcile_payments(report, bank)
    m1, m2, m3 = st.columns(3)
    m1.metric("Plăți potrivite", f"{rp['n_match']}")
    m2.metric("În raport, fără corespondent în bancă", f"{len(rp['rep_unmatched'])}")
    m3.metric("În bancă, fără corespondent în raport", f"{len(rp['bank_unmatched'])}")
    st.caption(f"Din {rp['n_rep']} plăți cu furnizor în raport și {rp['n_bank']} plăți către beneficiari în extras.")

    if len(rp["bank_unmatched"]):
        st.warning(f"**{len(rp['bank_unmatched'])} plăți există în bancă dar NU le găsesc în raport** "
                   "(bani ieșiți din cont, posibil neînregistrați sau cu altă sumă).")
        st.dataframe(rp["bank_unmatched"].style.format({"Sumă (lei)": "{:,.2f}"}),
                     use_container_width=True, hide_index=True)
        st.download_button("⬇️ Descarcă plățile din bancă fără corespondent (CSV)",
                           data=rp["bank_unmatched"].to_csv(index=False).encode("utf-8-sig"),
                           file_name="banca_fara_corespondent.csv", mime="text/csv", key="dl_bk_un")
    else:
        st.success("Toate plățile din bancă au corespondent în raport.")

    if len(rp["rep_unmatched"]):
        with st.expander(f"⚠️ {len(rp['rep_unmatched'])} rânduri din raport fără corespondent în bancă"):
            st.dataframe(rp["rep_unmatched"].style.format({"Sumă (lei)": "{:,.2f}"}),
                         use_container_width=True, hide_index=True)
            st.download_button("⬇️ Descarcă rândurile din raport fără corespondent (CSV)",
                               data=rp["rep_unmatched"].to_csv(index=False).encode("utf-8-sig"),
                               file_name="raport_fara_corespondent.csv", mime="text/csv", key="dl_rp_un")

    with st.expander(f"Vezi cele {rp['n_match']} plăți potrivite"):
        st.dataframe(rp["matches"].style.format(
            {c: "{:,.2f}" for c in ["Sumă (lei)", "Sold AG (raport)", "Sold real (bancă)", "Diferență sold"]
             if c in rp["matches"].columns}, na_rep="—"),
            use_container_width=True, hide_index=True)

    # ---- Verificarea soldului AG față de extras ----
    sc = rp.get("sold_check")
    if sc is not None and len(sc):
        st.divider()
        st.subheader("Verificarea soldului „Sold cont” (AG) față de extras")
        st.caption("Pentru fiecare plată potrivită cu extrasul, comparăm soldul scris în raport în coloana AG "
                   "cu soldul REAL din bancă imediat după acea operațiune. Așa vezi dacă soldul din raport "
                   "urmărește realitatea din cont.")
        cok, cbad = st.columns(2)
        cok.metric("Solduri AG care coincid cu banca", f"{rp['sold_ok']}")
        cbad.metric("Solduri AG diferite de bancă", f"{rp['sold_bad']}")
        if rp["sold_bad"] == 0:
            st.success("Soldul din raport (AG) coincide cu soldul real din bancă la toate plățile potrivite.")
        else:
            st.warning(f"**{rp['sold_bad']} solduri din raport (AG) NU coincid cu soldul real din bancă** "
                       "la momentul plății — soldul din raport nu reflectă realitatea din cont.")
            bad = sc[sc["Diferență sold"].abs() > 0.01]
            with st.expander(f"Vezi cele {len(bad)} solduri AG greșite față de bancă"):
                st.dataframe(bad.style.format(
                    {c: "{:,.2f}" for c in ["Sumă (lei)", "Sold AG (raport)", "Sold real (bancă)", "Diferență sold"]}),
                    use_container_width=True, hide_index=True)
            st.download_button("⬇️ Descarcă verificarea soldului AG vs extras (CSV)",
                               data=sc.to_csv(index=False).encode("utf-8-sig"),
                               file_name="sold_AG_vs_extras.csv", mime="text/csv", key="dl_ag")

    # ---- Raport de diferențe PE ZI ----
    st.divider()
    st.subheader("Diferențe pe zi (raport ↔ extras)")
    st.caption("Pentru fiecare zi calendaristică comparăm **intrările** (bani primiți: subvenție + venituri în "
               "clarificare/retur, în extras = creditele), **ieșirile** (cheltuieli + transferuri + comisioane, "
               "în extras = debitele) și **soldul la sfârșitul zilei**. Ieșirile includ de ambele părți comisioanele "
               "(în raport ca total zilnic pe «com. bancare», în extras ca linii de speze). Anul raportului este adus "
               "la anul extrasului, fiindcă în raport doar anul e greșit.")
    dd = daily_diff(report, bank)
    z1, z2, z3, z4 = st.columns(4)
    z1.metric("Zile analizate", f"{dd['n_zile']}")
    z2.metric("Zile cu diferențe la flux", f"{dd['n_zile_dif']}")
    z3.metric("Abatere maximă de sold", _fmt(dd["max_dif_sold"]))
    z4.metric("Diferență sold la final", _fmt(dd["final_dif_sold"]))
    st.caption("„Zile cu diferențe la flux” = zile în care intrările sau ieșirile diferă. Soldul este cumulativ: "
               "o dată apărută o diferență de flux, ea se reflectă în soldul tuturor zilelor următoare — de aceea "
               "îl arătăm separat (abaterea maximă și diferența finală).")
    if dd["n_zile_dif"] == 0:
        st.success("Nicio zi nu are diferențe între raport și extras la intrări/ieșiri.")
    else:
        st.warning(f"**{dd['n_zile_dif']} zile** au diferențe între raport și extras la intrări sau ieșiri.")
        fmt = {c: "{:,.2f}" for c in dd["bad"].columns if c != "Data"}
        with st.expander(f"Vezi cele {dd['n_zile_dif']} zile cu diferențe"):
            st.dataframe(dd["bad"].style.format(fmt).map(
                lambda x: "color:#c00;font-weight:600" if isinstance(x, (int, float)) and abs(x) > 0.01 else "",
                subset=["Dif. intrări", "Dif. ieșiri", "Dif. sold"]),
                use_container_width=True, hide_index=True)
    with st.expander(f"Vezi toate cele {dd['n_zile']} zile (cu sold pe zi)"):
        fmt2 = {c: "{:,.2f}" for c in dd["full"].columns if c != "Data"}
        st.dataframe(dd["full"].style.format(fmt2).map(
            lambda x: "color:#c00;font-weight:600" if isinstance(x, (int, float)) and abs(x) > 0.01 else "",
            subset=["Dif. intrări", "Dif. ieșiri", "Dif. sold"]),
            use_container_width=True, hide_index=True)
    st.download_button("⬇️ Descarcă raportul de diferențe pe zi (CSV)",
                       data=dd["full"].to_csv(index=False).encode("utf-8-sig"),
                       file_name="diferente_pe_zi.csv", mime="text/csv", key="dl_difzi")


def main():
    import streamlit as st

    st.set_page_config(page_title="Verificare subvenție", page_icon="📊", layout="wide")

    st.title("📊 Verificare subvenție — calcul independent")
    st.markdown(
        "Încarcă **raportul de subvenție** și/sau **extrasul de cont bancar** (.xls / .xlsx). "
        "Aplicația **recalculează singură totul, pas cu pas**, fără să se bazeze pe formulele din Excel "
        "(care pot fi greșite). Dacă încarci ambele fișiere, le și compară între ele."
    )

    ups = st.file_uploader(
        "Încarcă unul sau ambele fișiere",
        type=["xls", "xlsx"], accept_multiple_files=True,
    )
    if not ups:
        st.info("Aștept fișierele. Poți încărca raportul, extrasul, sau pe amândouă deodată.")
        st.stop()

    report_res, bank_res = None, None
    for up in ups:
        try:
            raw = read_excel_any(up, up.name)
            if is_bank_statement(raw):
                bank_res = analyze_bank(raw)
                st.success(f"„{up.name}” — recunoscut ca **extras de cont bancar**.")
            else:
                report_res = analyze(raw)
                st.success(f"„{up.name}” — recunoscut ca **raport de subvenție**.")
        except Exception as e:
            st.error(f"Nu am putut procesa „{up.name}”: {e}")

    if report_res is None and bank_res is None:
        st.stop()

    labels, renderers = [], []
    if bank_res is not None:
        labels.append("🏦 Extras de cont")
        renderers.append(lambda st=st: render_bank(st, bank_res))
    if report_res is not None:
        labels.append("📄 Raport subvenție")
        renderers.append(lambda st=st: render_report(st, report_res))
    if report_res is not None and bank_res is not None:
        labels.append("🔁 Reconciliere")
        renderers.append(lambda st=st: render_reconcile(st, report_res, bank_res))

    if len(labels) == 1:
        renderers[0]()
    else:
        for tab, render in zip(st.tabs(labels), renderers):
            with tab:
                render()


if __name__ == "__main__":
    main()